"""
거래명세표 Excel 자동 생성 스크립트
===================================
베어링/부품 거래명세표를 .xlsx로 생성한다.
코드번호 입력 시 품명/규격/단가가 VLOOKUP으로 자동 매핑되고,
수량 입력 시 공급가액이 자동 계산된다.

3부 구성: 공급자용(입력) → 공급받는자용(자동 복사) → 보관용(자동 복사)
1번째 양식에 입력하면 2, 3번째에 자동 반영된다.

사용법:
    python3 create_invoice.py

출력:
    거래명세표.xlsx (같은 디렉토리)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

# ──────────────────────────────────────────────
# 공통 스타일
# ──────────────────────────────────────────────
THIN = Side(style="thin")
MEDIUM = Side(style="medium")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MEDIUM_ALL = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FONT_TITLE = Font(name="맑은 고딕", size=18, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True)
FONT_NORMAL = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9)
FONT_TOTAL = Font(name="맑은 고딕", size=11, bold=True)
FONT_GRAND = Font(name="맑은 고딕", size=13, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

FILL_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FILL_HEADER = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

NUMBER_FMT = '#,##0'

# 양식 레이아웃 상수
FORM_ROWS = 23      # 한 양식의 행 수 (row 0~22 오프셋)
GAP_ROWS = 2        # 양식 간 빈 행 수
ITEM_COUNT = 9      # 거래 내역 행 수


def set_cell(ws, row, col, value, font=None, alignment=None, fill=None,
             number_format=None, border=None):
    """셀에 값과 스타일을 한번에 설정"""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    return cell


def ref(master_row, col):
    """마스터 셀 참조 수식 (빈 셀이면 빈 칸 표시)"""
    c = get_column_letter(col)
    return f'=IF({c}{master_row}="","",{c}{master_row})'


def apply_form_borders(ws, R):
    """양식 전체에 테두리를 빠짐없이 적용 (Row +0 ~ +19, Col A~I)"""
    for r in range(R, R + 20):  # 제목~합계(원정)까지
        for c in range(1, 10):  # A~I
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL

    # 서명란 하단 라인 (Row +21 ~ +22)
    for r in [R + 21, R + 22]:
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            # 서명란은 외곽만 얇은 하단선
            cell.border = Border(bottom=THIN)


# ──────────────────────────────────────────────
# 코드매핑 데이터 (~200개 품목)
# ──────────────────────────────────────────────
def get_code_mapping_data():
    """코드번호, 품명, 규격, 단가, 비고(용도) 데이터 반환"""
    data = [
        # === BEARING - 6000 시리즈 (소형 깊은홈) ===
        ("40040036401", "BEARING", "RNA6906", 5800, "112DT PUMP"),
        ("40040036402", "BEARING", "6000ZZ", 1200, "MOTOR"),
        ("40040036403", "BEARING", "6001ZZ", 1300, "MOTOR"),
        ("40040036404", "BEARING", "6002ZZ", 1400, "CONVEYOR"),
        ("40040036405", "BEARING", "6003ZZ", 1500, "CONVEYOR"),
        ("40040036406", "BEARING", "6004ZZ", 1800, "PUMP"),
        ("40040036407", "BEARING", "6005ZZ", 2200, "PUMP"),
        ("40040036408", "BEARING", "6006ZZ", 2800, "PUMP"),
        ("40040036409", "BEARING", "6007ZZ", 3500, "GEARBOX"),
        ("40040036410", "BEARING", "6008ZZ", 4200, "GEARBOX"),
        ("40040036411", "BEARING", "6009ZZ", 4800, "GEARBOX"),
        ("40040036412", "BEARING", "6010ZZ", 5500, "GEARBOX"),
        ("40040036413", "BEARING", "6011ZZ", 7800, "MOTOR"),
        ("40040036414", "BEARING", "6012ZZ", 9200, "MOTOR"),
        ("40040036415", "BEARING", "6013ZZ", 11500, "MOTOR"),

        # === BEARING - 6200 시리즈 (깊은홈) ===
        ("40040036416", "BEARING", "6200ZZ", 1500, "CONVEYOR"),
        ("40040036417", "BEARING", "6201ZZ", 1600, "CONVEYOR"),
        ("40040036418", "BEARING", "6202ZZ", 1700, "MOTOR"),
        ("40040036419", "BEARING", "6203ZZ", 1900, "MOTOR"),
        ("40040036420", "BEARING", "6204ZZ", 2100, "PUMP"),
        ("40040036421", "BEARING", "6205ZZ", 2500, "PUMP"),
        ("40040036422", "BEARING", "6206ZZ", 3200, "PUMP"),
        ("40040036423", "BEARING", "6207ZZ", 4000, "GEARBOX"),
        ("40040036424", "BEARING", "6208ZZ", 4800, "GEARBOX"),
        ("40040036425", "BEARING", "6209ZZ", 5600, "GEARBOX"),
        ("40040036426", "BEARING", "6210ZZ", 6500, "GEARBOX"),
        ("40040036427", "BEARING", "6211ZZ", 8500, "MOTOR"),
        ("40040036428", "BEARING", "6212ZZ", 10500, "MOTOR"),
        ("40040036429", "BEARING", "6213ZZ", 13000, "MOTOR"),
        ("40040036430", "BEARING", "6214ZZ", 15500, "MOTOR"),
        ("40040036431", "BEARING", "6215ZZ", 18000, "MOTOR"),

        # === BEARING - 6300 시리즈 (깊은홈, 대형) ===
        ("40040036432", "BEARING", "6300ZZ", 2000, "CONVEYOR"),
        ("40040036433", "BEARING", "6301ZZ", 2100, "CONVEYOR"),
        ("40040036434", "BEARING", "6302ZZ", 2300, "MOTOR"),
        ("40040036435", "BEARING", "6303ZZ", 2600, "MOTOR"),
        ("40040036436", "BEARING", "6304ZZ", 3000, "PUMP"),
        ("40040036437", "BEARING", "6305ZZ", 3500, "PUMP"),
        ("40040036438", "BEARING", "6306ZZ", 4200, "PUMP"),
        ("40040036439", "BEARING", "6307ZZ", 5500, "GEARBOX"),
        ("40040036440", "BEARING", "6308ZZ", 6800, "GEARBOX"),
        ("40040036441", "BEARING", "6309ZZ", 8200, "GEARBOX"),
        ("40040036442", "BEARING", "6310ZZ", 10000, "GEARBOX"),
        ("40040036443", "BEARING", "6311ZZ", 13000, "MOTOR"),
        ("40040036444", "BEARING", "6312ZZ", 16000, "MOTOR"),

        # === BEARING - 6800/6900 시리즈 (박형) ===
        ("40040036445", "BEARING", "6800ZZ", 1800, "SPINDLE"),
        ("40040036446", "BEARING", "6801ZZ", 1900, "SPINDLE"),
        ("40040036447", "BEARING", "6802ZZ", 2000, "SPINDLE"),
        ("40040036448", "BEARING", "6803ZZ", 2200, "SPINDLE"),
        ("40040036449", "BEARING", "6804ZZ", 2500, "SPINDLE"),
        ("40040036450", "BEARING", "6900ZZ", 1600, "SPINDLE"),
        ("40040036451", "BEARING", "6901ZZ", 1700, "SPINDLE"),
        ("40040036452", "BEARING", "6902ZZ", 1800, "SPINDLE"),
        ("40040036453", "BEARING", "6903ZZ", 2000, "SPINDLE"),
        ("40040036454", "BEARING", "6904ZZ", 2300, "SPINDLE"),
        ("40040036455", "BEARING", "6905ZZ", 2600, "SPINDLE"),

        # === BEARING - NUP/NU/NJ 시리즈 (원통 롤러) ===
        ("40040306101", "BEARING", "NUP308ET", 18500, "112DT PUMP"),
        ("40040306102", "BEARING", "NU204ET", 8500, "PUMP"),
        ("40040306103", "BEARING", "NU205ET", 9500, "PUMP"),
        ("40040306104", "BEARING", "NU206ET", 11000, "PUMP"),
        ("40040306105", "BEARING", "NU207ET", 13000, "GEARBOX"),
        ("40040306106", "BEARING", "NU208ET", 15000, "GEARBOX"),
        ("40040306107", "BEARING", "NU209ET", 17500, "GEARBOX"),
        ("40040306108", "BEARING", "NU210ET", 20000, "GEARBOX"),
        ("40040306109", "BEARING", "NU211ET", 23000, "MOTOR"),
        ("40040306110", "BEARING", "NU212ET", 27000, "MOTOR"),
        ("40040306111", "BEARING", "NJ204ET", 9000, "PUMP"),
        ("40040306112", "BEARING", "NJ205ET", 10000, "PUMP"),
        ("40040306113", "BEARING", "NJ206ET", 11500, "PUMP"),
        ("40040306114", "BEARING", "NJ207ET", 14000, "GEARBOX"),
        ("40040306115", "BEARING", "NJ208ET", 16000, "GEARBOX"),
        ("40040306116", "BEARING", "NUP204ET", 9500, "PUMP"),
        ("40040306117", "BEARING", "NUP205ET", 10500, "PUMP"),
        ("40040306118", "BEARING", "NUP206ET", 12000, "PUMP"),
        ("40040306119", "BEARING", "NUP305ET", 12000, "PUMP"),
        ("40040306120", "BEARING", "NUP306ET", 14000, "PUMP"),
        ("40040306121", "BEARING", "NUP307ET", 16500, "GEARBOX"),
        ("40040306122", "BEARING", "NUP309ET", 21000, "GEARBOX"),
        ("40040306123", "BEARING", "NUP310ET", 24000, "GEARBOX"),

        # === BEARING - RNA/NK/HK 니들 ===
        ("40040036460", "BEARING", "RNA4900", 3200, "GUIDE ROLLER"),
        ("40040036461", "BEARING", "RNA4901", 3500, "GUIDE ROLLER"),
        ("40040036462", "BEARING", "RNA4902", 3800, "GUIDE ROLLER"),
        ("40040036463", "BEARING", "RNA4903", 4200, "GUIDE ROLLER"),
        ("40040036464", "BEARING", "RNA4904", 4600, "GUIDE ROLLER"),
        ("40040036465", "BEARING", "RNA6901", 4800, "PUMP"),
        ("40040036466", "BEARING", "RNA6902", 5200, "PUMP"),
        ("40040036467", "BEARING", "RNA6903", 5500, "PUMP"),
        ("40040036468", "BEARING", "NK10/16", 2800, "CAM FOLLOWER"),
        ("40040036469", "BEARING", "NK12/16", 3000, "CAM FOLLOWER"),
        ("40040036470", "BEARING", "NK15/16", 3200, "CAM FOLLOWER"),
        ("40040036471", "BEARING", "NK17/16", 3500, "CAM FOLLOWER"),
        ("40040036472", "BEARING", "NK20/16", 3800, "CAM FOLLOWER"),
        ("40040036473", "BEARING", "HK0808", 1200, "CYLINDER"),
        ("40040036474", "BEARING", "HK1010", 1400, "CYLINDER"),
        ("40040036475", "BEARING", "HK1210", 1500, "CYLINDER"),
        ("40040036476", "BEARING", "HK1212", 1600, "CYLINDER"),
        ("40040036477", "BEARING", "HK1412", 1800, "CYLINDER"),
        ("40040036478", "BEARING", "HK1512", 1900, "CYLINDER"),
        ("40040036479", "BEARING", "HK1516", 2100, "CYLINDER"),
        ("40040036480", "BEARING", "HK2010", 2200, "CYLINDER"),

        # === BEARING - 자동조심 (22xx, 23xx) ===
        ("40040036481", "BEARING", "2204ETN9", 12000, "VIBRATION MOTOR"),
        ("40040036482", "BEARING", "2205ETN9", 13500, "VIBRATION MOTOR"),
        ("40040036483", "BEARING", "2206ETN9", 15000, "VIBRATION MOTOR"),
        ("40040036484", "BEARING", "2207ETN9", 17000, "VIBRATION MOTOR"),
        ("40040036485", "BEARING", "2208ETN9", 19000, "VIBRATION MOTOR"),
        ("40040036486", "BEARING", "2209ETN9", 22000, "VIBRATION MOTOR"),
        ("40040036487", "BEARING", "2210ETN9", 25000, "VIBRATION MOTOR"),
        ("40040036488", "BEARING", "22205E", 18000, "HEAVY EQUIPMENT"),
        ("40040036489", "BEARING", "22206E", 20000, "HEAVY EQUIPMENT"),
        ("40040036490", "BEARING", "22207E", 23000, "HEAVY EQUIPMENT"),
        ("40040036491", "BEARING", "22208E", 26000, "HEAVY EQUIPMENT"),
        ("40040036492", "BEARING", "22209E", 29000, "HEAVY EQUIPMENT"),
        ("40040036493", "BEARING", "22210E", 32000, "HEAVY EQUIPMENT"),
        ("40040036494", "BEARING", "22211E", 36000, "HEAVY EQUIPMENT"),
        ("40040036495", "BEARING", "22212E", 40000, "HEAVY EQUIPMENT"),
        ("40040036496", "BEARING", "23024E", 85000, "CRUSHER"),
        ("40040036497", "BEARING", "23026E", 95000, "CRUSHER"),
        ("40040036498", "BEARING", "23028E", 110000, "CRUSHER"),
        ("40040036499", "BEARING", "23030E", 125000, "CRUSHER"),

        # === BEARING - 앵귤러 컨택트 ===
        ("40040036500", "BEARING", "7200B", 4500, "SPINDLE"),
        ("40040036501", "BEARING", "7201B", 4800, "SPINDLE"),
        ("40040036502", "BEARING", "7202B", 5200, "SPINDLE"),
        ("40040036503", "BEARING", "7203B", 5600, "SPINDLE"),
        ("40040036504", "BEARING", "7204B", 6200, "SPINDLE"),
        ("40040036505", "BEARING", "7205B", 7000, "SPINDLE"),
        ("40040036506", "BEARING", "7206B", 8000, "SPINDLE"),
        ("40040036507", "BEARING", "7207B", 9200, "SPINDLE"),
        ("40040036508", "BEARING", "7208B", 10500, "SPINDLE"),

        # === BEARING - 테이퍼 롤러 ===
        ("40040036510", "BEARING", "30204", 5000, "WHEEL"),
        ("40040036511", "BEARING", "30205", 5500, "WHEEL"),
        ("40040036512", "BEARING", "30206", 6200, "WHEEL"),
        ("40040036513", "BEARING", "30207", 7000, "WHEEL"),
        ("40040036514", "BEARING", "30208", 8000, "WHEEL"),
        ("40040036515", "BEARING", "30209", 9000, "WHEEL"),
        ("40040036516", "BEARING", "30210", 10000, "WHEEL"),
        ("40040036517", "BEARING", "32204", 6500, "GEARBOX"),
        ("40040036518", "BEARING", "32205", 7200, "GEARBOX"),
        ("40040036519", "BEARING", "32206", 8000, "GEARBOX"),
        ("40040036520", "BEARING", "32207", 9200, "GEARBOX"),
        ("40040036521", "BEARING", "32208", 10500, "GEARBOX"),

        # === OIL SEAL - TC 타입 ===
        ("55040036601", "OIL SEAL", "TC 20×35×7", 1800, "PUMP SHAFT"),
        ("55040036602", "OIL SEAL", "TC 25×40×7", 2000, "PUMP SHAFT"),
        ("55040036603", "OIL SEAL", "TC 25×42×7", 2100, "MOTOR SHAFT"),
        ("55040036604", "OIL SEAL", "TC 30×45×8", 2300, "MOTOR SHAFT"),
        ("55040036605", "OIL SEAL", "TC 30×47×7", 2400, "GEARBOX"),
        ("55040036606", "OIL SEAL", "TC 32×45×8", 2500, "GEARBOX"),
        ("55040036607", "OIL SEAL", "TC 35×50×8", 2800, "GEARBOX"),
        ("55040036608", "OIL SEAL", "TC 35×52×8", 2900, "PUMP"),
        ("55040036609", "OIL SEAL", "TC 38×55×8", 3200, "PUMP"),
        ("55040036610", "OIL SEAL", "TC 40×55×8", 3300, "PUMP"),
        ("55040036611", "OIL SEAL", "TC 40×58×8", 3500, "MOTOR"),
        ("55040036612", "OIL SEAL", "TC 42×60×8", 3700, "MOTOR"),
        ("55040036613", "OIL SEAL", "TC 45×62×8", 3900, "CYLINDER"),
        ("55040036614", "OIL SEAL", "TC 45×65×10", 4200, "CYLINDER"),
        ("55040036615", "OIL SEAL", "TC 48×65×10", 4400, "CYLINDER"),

        # === OIL SEAL - SC/VC 타입 ===
        ("55040036616", "OIL SEAL", "SC 50×68×10", 4600, "PUMP"),
        ("55040036617", "OIL SEAL", "SC 50×72×10", 4900, "PUMP"),
        ("55040036618", "OIL SEAL", "SC 55×72×10", 5200, "GEARBOX"),
        ("55040036619", "OIL SEAL", "SC 55×75×10", 5500, "GEARBOX"),
        ("55040036620", "OIL SEAL", "SC 60×80×10", 5800, "MOTOR"),
        ("55040036621", "OIL SEAL", "SC 60×82×12", 6200, "MOTOR"),
        ("55040036622", "OIL SEAL", "SC 65×85×10", 6500, "MOTOR"),
        ("55040036623", "OIL SEAL", "SC 70×90×10", 7000, "REDUCER"),
        ("55040036624", "OIL SEAL", "SC 75×95×12", 7500, "REDUCER"),
        ("55040036625", "OIL SEAL", "SC 80×100×12", 8000, "REDUCER"),
        ("55040036626", "OIL SEAL", "VC 25×35×6", 2200, "VALVE"),
        ("55040036627", "OIL SEAL", "VC 30×40×7", 2500, "VALVE"),
        ("55040036628", "OIL SEAL", "VC 35×47×7", 2800, "VALVE"),
        ("55040036629", "OIL SEAL", "VC 40×52×8", 3200, "HYDRAULIC"),
        ("55040036630", "OIL SEAL", "VC 45×60×8", 3600, "HYDRAULIC"),

        # === COLLAR - IR 시리즈 이너링 ===
        ("33040036701", "COLLAR", "IR 10×14×13", 2500, "NEEDLE BEARING"),
        ("33040036702", "COLLAR", "IR 12×15×13", 2600, "NEEDLE BEARING"),
        ("33040036703", "COLLAR", "IR 12×16×13", 2700, "NEEDLE BEARING"),
        ("33040036704", "COLLAR", "IR 15×18×16", 2900, "NEEDLE BEARING"),
        ("33040036705", "COLLAR", "IR 15×19×16", 3000, "NEEDLE BEARING"),
        ("33040036706", "COLLAR", "IR 17×20×16", 3200, "NEEDLE BEARING"),
        ("33040036707", "COLLAR", "IR 17×21×16", 3300, "NEEDLE BEARING"),
        ("33040036708", "COLLAR", "IR 20×24×16", 3500, "NEEDLE BEARING"),
        ("33040036709", "COLLAR", "IR 20×25×20", 3800, "NEEDLE BEARING"),
        ("33040036710", "COLLAR", "IR 22×26×16", 3600, "NEEDLE BEARING"),
        ("33040036711", "COLLAR", "IR 25×29×16", 3900, "NEEDLE BEARING"),
        ("33040036712", "COLLAR", "IR 25×30×20", 4200, "NEEDLE BEARING"),
        ("33040036713", "COLLAR", "IR 28×32×17", 4000, "NEEDLE BEARING"),
        ("33040036714", "COLLAR", "IR 30×34×17", 4300, "NEEDLE BEARING"),
        ("33040036715", "COLLAR", "IR 30×35×20", 4600, "NEEDLE BEARING"),
        ("33040036716", "COLLAR", "IR 32×37×17", 4500, "NEEDLE BEARING"),
        ("33040036717", "COLLAR", "IR 35×40×20", 5000, "NEEDLE BEARING"),
        ("33040036718", "COLLAR", "IR 40×45×20", 5500, "NEEDLE BEARING"),
        ("33040036719", "COLLAR", "IR 45×50×25", 6200, "NEEDLE BEARING"),
        ("33040036720", "COLLAR", "IR 50×55×25", 6800, "NEEDLE BEARING"),

        # === STEEL BALL ===
        ("40040036801", "STEEL BALL", '1/4" (6.35mm)', 50, "VALVE"),
        ("40040036802", "STEEL BALL", '5/16" (7.94mm)', 60, "VALVE"),
        ("40040036803", "STEEL BALL", '3/8" (9.53mm)', 80, "PUMP CHECK"),
        ("40040036804", "STEEL BALL", '7/16" (11.11mm)', 100, "PUMP CHECK"),
        ("40040036805", "STEEL BALL", '1/2" (12.7mm)', 120, "CHECK VALVE"),
        ("40040036806", "STEEL BALL", '5/8" (15.88mm)', 180, "CHECK VALVE"),
        ("40040036807", "STEEL BALL", '3/4" (19.05mm)', 250, "REGULATOR"),
        ("40040036808", "STEEL BALL", '7/8" (22.23mm)', 350, "REGULATOR"),
        ("40040036809", "STEEL BALL", '1" (25.4mm)', 500, "REGULATOR"),
        ("40040036810", "STEEL BALL", '1-1/4" (31.75mm)', 800, "SPECIAL"),

        # === O-RING ===
        ("66040036901", "O-RING", "P-6 (5.8×1.9)", 200, "CYLINDER"),
        ("66040036902", "O-RING", "P-8 (7.8×1.9)", 200, "CYLINDER"),
        ("66040036903", "O-RING", "P-10 (9.8×2.4)", 250, "CYLINDER"),
        ("66040036904", "O-RING", "P-12 (11.8×2.4)", 250, "CYLINDER"),
        ("66040036905", "O-RING", "P-14 (13.8×2.4)", 300, "PUMP"),
        ("66040036906", "O-RING", "P-16 (15.8×2.4)", 300, "PUMP"),
        ("66040036907", "O-RING", "G-25 (24.4×3.1)", 350, "PIPE JOINT"),
        ("66040036908", "O-RING", "G-30 (29.4×3.1)", 400, "PIPE JOINT"),
        ("66040036909", "O-RING", "G-35 (34.4×3.1)", 450, "PIPE JOINT"),
        ("66040036910", "O-RING", "G-40 (39.4×3.1)", 500, "FLANGE"),
        ("66040036911", "O-RING", "S-30 (29.7×2.4)", 400, "SHAFT"),
        ("66040036912", "O-RING", "S-35 (34.7×2.4)", 450, "SHAFT"),

        # === BUSH ===
        ("77040037001", "BUSH", "DU 1010", 1500, "GUIDE"),
        ("77040037002", "BUSH", "DU 1210", 1600, "GUIDE"),
        ("77040037003", "BUSH", "DU 1510", 1800, "GUIDE"),
        ("77040037004", "BUSH", "DU 1512", 2000, "GUIDE"),
        ("77040037005", "BUSH", "DU 2010", 2200, "SLIDE"),
        ("77040037006", "BUSH", "DU 2012", 2400, "SLIDE"),
        ("77040037007", "BUSH", "DU 2015", 2600, "SLIDE"),
        ("77040037008", "BUSH", "DU 2510", 2800, "SLIDE"),
        ("77040037009", "BUSH", "DU 2515", 3200, "PIVOT"),
        ("77040037010", "BUSH", "DU 3010", 3500, "PIVOT"),
    ]
    return data


# ──────────────────────────────────────────────
# Sheet 2: 코드매핑 시트 생성
# ──────────────────────────────────────────────
def create_code_mapping_sheet(wb):
    ws = wb.create_sheet("코드매핑")
    data = get_code_mapping_data()

    headers = ["코드번호", "품명", "규격", "단가", "비고"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, 1, col, header, font=FONT_HEADER, alignment=ALIGN_CENTER,
                 fill=FILL_HEADER, border=BORDER_ALL)

    for i, (code, name, spec, price, note) in enumerate(data, 2):
        set_cell(ws, i, 1, code, font=FONT_NORMAL, alignment=ALIGN_CENTER,
                 border=BORDER_ALL)
        set_cell(ws, i, 2, name, font=FONT_NORMAL, alignment=ALIGN_CENTER,
                 border=BORDER_ALL)
        set_cell(ws, i, 3, spec, font=FONT_NORMAL, alignment=ALIGN_CENTER,
                 border=BORDER_ALL)
        set_cell(ws, i, 4, price, font=FONT_NORMAL, alignment=ALIGN_RIGHT,
                 number_format=NUMBER_FMT, border=BORDER_ALL)
        set_cell(ws, i, 5, note, font=FONT_NORMAL, alignment=ALIGN_CENTER,
                 border=BORDER_ALL)

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    return len(data)


# ──────────────────────────────────────────────
# 셀 병합 구조 정의 (상대 행 기준)
# ──────────────────────────────────────────────
def get_merge_ranges():
    """양식의 셀 병합 범위를 (상대행시작, 상대행끝, 열시작, 열끝) 형태로 반환."""
    return [
        # Row 0: 제목
        (0, 0, 1, 9),
        # Row 1: 날짜
        (1, 1, 1, 9),
        # Row 2~6: 공급받는자
        (2, 6, 1, 1),      # A 세로 병합
        (2, 2, 2, 4),      # B3:D3
        (3, 3, 3, 4),      # C4:D4
        (4, 4, 2, 4),      # B5:D5
        (5, 5, 2, 4),      # B6:D6
        (6, 6, 2, 4),      # B7:D7
        # Row 2~6: 공급자
        (2, 6, 5, 5),      # E 세로 병합
        (2, 2, 6, 9),      # F3:I3
        (3, 3, 6, 7),      # F4:G4
        (3, 3, 8, 9),      # H4:I4
        (4, 4, 6, 9),      # F5:I5
        (5, 5, 6, 9),      # F6:I6
        (6, 6, 6, 9),      # F7:I7
        # Row 17~19: 합계
        (17, 17, 1, 7),
        (18, 18, 1, 7),
        (19, 19, 1, 7),
        # Row 21~22: 서명란
        (21, 21, 1, 4),
        (21, 21, 6, 9),
        (22, 22, 1, 4),
    ]


def apply_merges(ws, start_row, merge_ranges):
    """시작행 기준으로 셀 병합 적용"""
    for r1, r2, c1, c2 in merge_ranges:
        ws.merge_cells(
            start_row=start_row + r1,
            end_row=start_row + r2,
            start_column=c1,
            end_column=c2
        )


# ──────────────────────────────────────────────
# 마스터 양식 생성 (1번째 - 입력용)
# ──────────────────────────────────────────────
def build_master_form(ws, R, data_count, label="공급자용"):
    """R = 시작행 (1-based)."""
    merge_ranges = get_merge_ranges()
    apply_merges(ws, R, merge_ranges)

    # Row heights
    row_heights = {
        0: 35, 1: 22,
        2: 20, 3: 20, 4: 20, 5: 20, 6: 20,
        7: 25,
        17: 25, 18: 25, 19: 28,
        21: 20, 22: 20,
    }
    for offset in range(ITEM_COUNT):
        row_heights[8 + offset] = 22
    for offset, height in row_heights.items():
        ws.row_dimensions[R + offset].height = height

    # Row 0: 제목
    set_cell(ws, R, 1, f"거 래 명 세 표 ({label})",
             font=FONT_TITLE, alignment=ALIGN_CENTER)

    # Row 1: 날짜 (입력 셀)
    set_cell(ws, R+1, 1, "날짜:                    년      월      일",
             font=FONT_NORMAL, alignment=ALIGN_RIGHT, fill=FILL_YELLOW)

    # Row 2~6: 공급받는자
    set_cell(ws, R+2, 1, "공\n급\n받\n는\n자",
             font=FONT_HEADER, alignment=ALIGN_CENTER)
    set_cell(ws, R+2, 2, "등록번호: 603-81-70100",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 2, "상호: ㈜제일피엠씨",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 3, "성명:",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+4, 2, "주소: 경남 창원시 성산구 창원대로 885",
             font=FONT_SMALL, alignment=ALIGN_LEFT)
    set_cell(ws, R+5, 2, "업태: 제조업",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+6, 2, "종목: 산업기계 부품",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row 2~6: 공급자
    set_cell(ws, R+2, 5, "공\n급\n자",
             font=FONT_HEADER, alignment=ALIGN_CENTER)
    set_cell(ws, R+2, 6, "등록번호: 605-81-10092",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 6, "상호: ㈜동명베아링",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 8, "성명:          (인)",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+4, 6, "주소: 경남 창원시 의창구 중앙대로 183",
             font=FONT_SMALL, alignment=ALIGN_LEFT)
    set_cell(ws, R+5, 6, "업태: 도소매업",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+6, 6, "종목: 베어링, 산업용품",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row 7: 테이블 헤더
    table_headers = ["NO", "코드번호", "품명", "규격", "수량", "단위",
                     "단가", "공급가액", "비고"]
    for col, header in enumerate(table_headers, 1):
        set_cell(ws, R+7, col, header, font=FONT_HEADER, alignment=ALIGN_CENTER,
                 fill=FILL_HEADER)

    # Row 8~16: 거래 내역 (9행)
    for i in range(ITEM_COUNT):
        row = R + 8 + i
        code_cell = f"B{row}"
        qty = f"E{row}"
        price = f"G{row}"

        set_cell(ws, row, 1, i + 1,
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        set_cell(ws, row, 2, None,
                 font=FONT_NORMAL, alignment=ALIGN_CENTER, fill=FILL_YELLOW)
        set_cell(ws, row, 3,
                 f'=IFERROR(VLOOKUP({code_cell},코드매핑!$A:$E,2,FALSE),"")',
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        set_cell(ws, row, 4,
                 f'=IFERROR(VLOOKUP({code_cell},코드매핑!$A:$E,3,FALSE),"")',
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        set_cell(ws, row, 5, None,
                 font=FONT_NORMAL, alignment=ALIGN_CENTER, fill=FILL_YELLOW)
        set_cell(ws, row, 6, f'=IF({code_cell}="","","EA")',
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        set_cell(ws, row, 7,
                 f'=IFERROR(VLOOKUP({code_cell},코드매핑!$A:$E,4,FALSE),"")',
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT)
        set_cell(ws, row, 8,
                 f'=IF(AND({qty}<>"",{price}<>""),{qty}*{price},"")',
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT)
        set_cell(ws, row, 9,
                 f'=IFERROR(VLOOKUP({code_cell},코드매핑!$A:$E,5,FALSE),"")',
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)

    ITEM_START = R + 8
    ITEM_END = R + 8 + ITEM_COUNT - 1
    TOTAL_ROW = R + 17
    TAX_ROW = R + 18
    GRAND_ROW = R + 19

    # Row 17: 공급가액 합계
    set_cell(ws, TOTAL_ROW, 1, "공급가액 합계",
             font=FONT_TOTAL, alignment=ALIGN_CENTER, fill=FILL_LIGHT_GRAY)
    set_cell(ws, TOTAL_ROW, 8, f'=SUM(H{ITEM_START}:H{ITEM_END})',
             font=FONT_TOTAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT,
             fill=FILL_LIGHT_GRAY)
    set_cell(ws, TOTAL_ROW, 9, "", font=FONT_NORMAL, alignment=ALIGN_CENTER,
             fill=FILL_LIGHT_GRAY)
    # 합계 행 나머지 셀 fill
    for c in range(2, 8):
        ws.cell(row=TOTAL_ROW, column=c).fill = FILL_LIGHT_GRAY

    # Row 18: 세액
    set_cell(ws, TAX_ROW, 1, "세액 (10%)",
             font=FONT_TOTAL, alignment=ALIGN_CENTER)
    set_cell(ws, TAX_ROW, 8, f'=IF(H{TOTAL_ROW}=0,"",INT(H{TOTAL_ROW}*0.1))',
             font=FONT_TOTAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT)
    set_cell(ws, TAX_ROW, 9, "", font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Row 19: 합계(원정)
    set_cell(ws, GRAND_ROW, 1, "합 계 (원정)",
             font=FONT_GRAND, alignment=ALIGN_CENTER, fill=FILL_HEADER)
    set_cell(ws, GRAND_ROW, 8,
             f'=IF(H{TOTAL_ROW}=0,"",H{TOTAL_ROW}+H{TAX_ROW})',
             font=FONT_GRAND, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT,
             fill=FILL_HEADER)
    set_cell(ws, GRAND_ROW, 9, "", font=FONT_NORMAL, alignment=ALIGN_CENTER,
             fill=FILL_HEADER)
    for c in range(2, 8):
        ws.cell(row=GRAND_ROW, column=c).fill = FILL_HEADER

    # Row 21: 서명란
    set_cell(ws, R+21, 1, "인수자:                    (인)",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+21, 6, "품질관리:                  (인)",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row 22: 인수일자
    set_cell(ws, R+22, 1, "인수일자:      년    월    일",
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # 전체 테두리 적용 (누락 방지)
    apply_form_borders(ws, R)

    # 데이터 유효성 검사 (코드번호 드롭다운)
    code_range = f"코드매핑!$A$2:$A${data_count + 1}"
    dv = DataValidation(
        type="list",
        formula1=code_range,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="코드번호 오류",
        error="코드매핑 시트에 등록된 코드번호를 선택하세요."
    )
    dv.prompt = "코드번호를 선택하세요"
    dv.promptTitle = "코드번호"
    dv.showInputMessage = True
    for i in range(ITEM_COUNT):
        dv.add(f"B{ITEM_START + i}")
    ws.add_data_validation(dv)


# ──────────────────────────────────────────────
# 미러 양식 생성 (2, 3번째 - 자동 복사)
# ──────────────────────────────────────────────
def build_mirror_form(ws, R, master_R, label="공급받는자용"):
    """
    마스터 양식을 참조하는 미러 양식 생성.
    R = 미러 시작행, master_R = 마스터 시작행.
    """
    merge_ranges = get_merge_ranges()
    apply_merges(ws, R, merge_ranges)

    # Row heights
    row_heights = {
        0: 35, 1: 22,
        2: 20, 3: 20, 4: 20, 5: 20, 6: 20,
        7: 25,
        17: 25, 18: 25, 19: 28,
        21: 20, 22: 20,
    }
    for offset in range(ITEM_COUNT):
        row_heights[8 + offset] = 22
    for offset, height in row_heights.items():
        ws.row_dimensions[R + offset].height = height

    # Row 0: 제목 (유일하게 다른 텍스트)
    set_cell(ws, R, 1, f"거 래 명 세 표 ({label})",
             font=FONT_TITLE, alignment=ALIGN_CENTER)

    # Row 1: 날짜 → 마스터 참조
    set_cell(ws, R+1, 1, ref(master_R+1, 1),
             font=FONT_NORMAL, alignment=ALIGN_RIGHT)

    # Row 2~6: 공급받는자 → 마스터 참조
    set_cell(ws, R+2, 1, ref(master_R+2, 1),
             font=FONT_HEADER, alignment=ALIGN_CENTER)
    set_cell(ws, R+2, 2, ref(master_R+2, 2),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 2, ref(master_R+3, 2),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 3, ref(master_R+3, 3),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+4, 2, ref(master_R+4, 2),
             font=FONT_SMALL, alignment=ALIGN_LEFT)
    set_cell(ws, R+5, 2, ref(master_R+5, 2),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+6, 2, ref(master_R+6, 2),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row 2~6: 공급자 → 마스터 참조
    set_cell(ws, R+2, 5, ref(master_R+2, 5),
             font=FONT_HEADER, alignment=ALIGN_CENTER)
    set_cell(ws, R+2, 6, ref(master_R+2, 6),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 6, ref(master_R+3, 6),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+3, 8, ref(master_R+3, 8),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+4, 6, ref(master_R+4, 6),
             font=FONT_SMALL, alignment=ALIGN_LEFT)
    set_cell(ws, R+5, 6, ref(master_R+5, 6),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+6, 6, ref(master_R+6, 6),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row 7: 테이블 헤더 → 마스터 참조 + 색상 동일
    for col in range(1, 10):
        set_cell(ws, R+7, col, ref(master_R+7, col),
                 font=FONT_HEADER, alignment=ALIGN_CENTER, fill=FILL_HEADER)

    # Row 8~16: 거래 내역 → 마스터 참조 (빈 셀 = 빈 칸)
    for i in range(ITEM_COUNT):
        row = R + 8 + i
        master_row = master_R + 8 + i

        # NO (항상 값 있음)
        set_cell(ws, row, 1, ref(master_row, 1),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        # 코드번호
        set_cell(ws, row, 2, ref(master_row, 2),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        # 품명
        set_cell(ws, row, 3, ref(master_row, 3),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        # 규격
        set_cell(ws, row, 4, ref(master_row, 4),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        # 수량
        set_cell(ws, row, 5, ref(master_row, 5),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        # 단위
        set_cell(ws, row, 6, ref(master_row, 6),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)
        # 단가
        set_cell(ws, row, 7, ref(master_row, 7),
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT)
        # 공급가액
        set_cell(ws, row, 8, ref(master_row, 8),
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT)
        # 비고
        set_cell(ws, row, 9, ref(master_row, 9),
                 font=FONT_NORMAL, alignment=ALIGN_CENTER)

    TOTAL_ROW = R + 17
    TAX_ROW = R + 18
    GRAND_ROW = R + 19
    M_TOTAL = master_R + 17
    M_TAX = master_R + 18
    M_GRAND = master_R + 19

    # Row 17: 공급가액 합계 → 색상 + 마스터 참조
    set_cell(ws, TOTAL_ROW, 1, ref(M_TOTAL, 1),
             font=FONT_TOTAL, alignment=ALIGN_CENTER, fill=FILL_LIGHT_GRAY)
    set_cell(ws, TOTAL_ROW, 8, ref(M_TOTAL, 8),
             font=FONT_TOTAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT,
             fill=FILL_LIGHT_GRAY)
    set_cell(ws, TOTAL_ROW, 9, "", font=FONT_NORMAL, alignment=ALIGN_CENTER,
             fill=FILL_LIGHT_GRAY)
    for c in range(2, 8):
        ws.cell(row=TOTAL_ROW, column=c).fill = FILL_LIGHT_GRAY

    # Row 18: 세액 → 마스터 참조
    set_cell(ws, TAX_ROW, 1, ref(M_TAX, 1),
             font=FONT_TOTAL, alignment=ALIGN_CENTER)
    set_cell(ws, TAX_ROW, 8, ref(M_TAX, 8),
             font=FONT_TOTAL, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT)
    set_cell(ws, TAX_ROW, 9, "", font=FONT_NORMAL, alignment=ALIGN_CENTER)

    # Row 19: 합계(원정) → 색상 + 마스터 참조
    set_cell(ws, GRAND_ROW, 1, ref(M_GRAND, 1),
             font=FONT_GRAND, alignment=ALIGN_CENTER, fill=FILL_HEADER)
    set_cell(ws, GRAND_ROW, 8, ref(M_GRAND, 8),
             font=FONT_GRAND, alignment=ALIGN_RIGHT, number_format=NUMBER_FMT,
             fill=FILL_HEADER)
    set_cell(ws, GRAND_ROW, 9, "", font=FONT_NORMAL, alignment=ALIGN_CENTER,
             fill=FILL_HEADER)
    for c in range(2, 8):
        ws.cell(row=GRAND_ROW, column=c).fill = FILL_HEADER

    # Row 21: 서명란 → 마스터 참조
    set_cell(ws, R+21, 1, ref(master_R+21, 1),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)
    set_cell(ws, R+21, 6, ref(master_R+21, 6),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row 22: 인수일자 → 마스터 참조
    set_cell(ws, R+22, 1, ref(master_R+22, 1),
             font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # 전체 테두리 적용 (누락 방지)
    apply_form_borders(ws, R)


# ──────────────────────────────────────────────
# Sheet 1: 거래명세표 시트 (3부 구성)
# ──────────────────────────────────────────────
def create_invoice_sheet(wb, data_count):
    ws = wb.active
    ws.title = "거래명세표"

    # 열 너비 설정 (이미지 양식 비율에 맞게 조정)
    col_widths = {
        'A': 4.5,   # NO
        'B': 17,    # 코드번호 (11자리)
        'C': 12,    # 품명
        'D': 17,    # 규격 (TC 20×35×7 등)
        'E': 7,     # 수량
        'F': 5.5,   # 단위
        'G': 11,    # 단가
        'H': 13,    # 공급가액
        'I': 13,    # 비고
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Form 1: 공급자용 (마스터 - 입력)
    MASTER_START = 1
    build_master_form(ws, MASTER_START, data_count, label="공급자용")

    # Form 2: 공급받는자용 (미러)
    MIRROR2_START = MASTER_START + FORM_ROWS + GAP_ROWS  # 26
    build_mirror_form(ws, MIRROR2_START, MASTER_START, label="공급받는자용")

    # Form 3: 보관용 (미러)
    MIRROR3_START = MIRROR2_START + FORM_ROWS + GAP_ROWS  # 51
    build_mirror_form(ws, MIRROR3_START, MASTER_START, label="보관용")

    LAST_ROW = MIRROR3_START + FORM_ROWS - 1  # 73

    # 인쇄 설정 (A4 가로, 3페이지)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 높이 자동 (3페이지)
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )

    ws.print_area = f"A1:I{LAST_ROW}"

    # 페이지 나누기
    ws.row_breaks.append(Break(id=MIRROR2_START - 1))
    ws.row_breaks.append(Break(id=MIRROR3_START - 1))

    return ws


# ──────────────────────────────────────────────
# 메인 실행
# ──────────────────────────────────────────────
def main():
    wb = Workbook()

    data_count = create_code_mapping_sheet(wb)
    print(f"코드매핑 시트 생성 완료: {data_count}개 품목")

    create_invoice_sheet(wb, data_count)
    print("거래명세표 시트 생성 완료 (3부: 공급자용 / 공급받는자용 / 보관용)")

    wb.move_sheet("거래명세표", offset=-1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "거래명세표.xlsx")
    wb.save(output_path)
    print(f"\n저장 완료: {output_path}")
    print(f"총 {data_count}개 품목이 등록되었습니다.")
    print("\n구성:")
    print("  1부 (Row 1~23):  공급자용 - 여기에 입력")
    print("  2부 (Row 26~48): 공급받는자용 - 자동 복사")
    print("  3부 (Row 51~73): 보관용 - 자동 복사")
    print("\n1부에 코드번호/수량을 입력하면 2부, 3부에 자동 반영됩니다.")


if __name__ == "__main__":
    main()
